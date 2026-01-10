"""
Generate test XLS files for Qualys and Tenable vulnerability reports.

Creates 3 weeks of data for each scanner:
- Week 1: 100 vulnerabilities (baseline)
- Week 2: 90 from week 1 + 10 new (10% new)
- Week 3: 72 from week 1 + 10 from week 2 + 5 new (20% of original closed, 5% new)
"""

import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
import random

# Common CVEs and vulnerability data
CVES = [
    ("CVE-2024-21762", "Fortinet FortiOS Out-of-Bounds Write", 9.8, "critical"),
    ("CVE-2024-1709", "ConnectWise ScreenConnect Authentication Bypass", 10.0, "critical"),
    ("CVE-2024-27198", "JetBrains TeamCity Authentication Bypass", 9.8, "critical"),
    ("CVE-2024-3400", "Palo Alto PAN-OS Command Injection", 10.0, "critical"),
    ("CVE-2024-21887", "Ivanti Connect Secure Command Injection", 9.1, "critical"),
    ("CVE-2023-46805", "Ivanti Connect Secure Authentication Bypass", 8.2, "high"),
    ("CVE-2024-23897", "Jenkins Arbitrary File Read", 9.8, "critical"),
    ("CVE-2024-0204", "GoAnywhere MFT Authentication Bypass", 9.8, "critical"),
    ("CVE-2023-22527", "Atlassian Confluence RCE", 9.8, "critical"),
    ("CVE-2024-1212", "Progress LoadMaster Command Injection", 10.0, "critical"),
    ("CVE-2023-4966", "Citrix NetScaler Information Disclosure", 9.4, "critical"),
    ("CVE-2023-20198", "Cisco IOS XE Web UI Privilege Escalation", 10.0, "critical"),
    ("CVE-2023-42793", "JetBrains TeamCity Authentication Bypass", 9.8, "critical"),
    ("CVE-2023-29357", "Microsoft SharePoint Privilege Escalation", 9.8, "critical"),
    ("CVE-2023-35078", "Ivanti EPMM Authentication Bypass", 10.0, "critical"),
    ("CVE-2023-27997", "Fortinet FortiOS Heap Buffer Overflow", 9.8, "critical"),
    ("CVE-2023-34362", "MOVEit Transfer SQL Injection", 9.8, "critical"),
    ("CVE-2023-2868", "Barracuda ESG Command Injection", 9.8, "critical"),
    ("CVE-2022-47966", "Zoho ManageEngine RCE", 9.8, "critical"),
    ("CVE-2022-41040", "Microsoft Exchange SSRF", 8.8, "high"),
    ("CVE-2022-26134", "Atlassian Confluence OGNL Injection", 9.8, "critical"),
    ("CVE-2022-22965", "Spring Framework RCE (Spring4Shell)", 9.8, "critical"),
    ("CVE-2021-44228", "Apache Log4j RCE (Log4Shell)", 10.0, "critical"),
    ("CVE-2021-34473", "Microsoft Exchange Server RCE", 9.8, "critical"),
    ("CVE-2021-26855", "Microsoft Exchange Server SSRF (ProxyLogon)", 9.8, "critical"),
    ("CVE-2023-38831", "WinRAR Code Execution", 7.8, "high"),
    ("CVE-2023-23397", "Microsoft Outlook Privilege Escalation", 9.8, "critical"),
    ("CVE-2023-36884", "Microsoft Office RCE", 8.3, "high"),
    ("CVE-2023-28252", "Windows CLFS Privilege Escalation", 7.8, "high"),
    ("CVE-2023-21823", "Windows Graphics Component RCE", 7.8, "high"),
    ("CVE-2022-30190", "Microsoft Windows Support Diagnostic Tool RCE", 7.8, "high"),
    ("CVE-2022-41082", "Microsoft Exchange Server RCE", 8.8, "high"),
    ("CVE-2023-44487", "HTTP/2 Rapid Reset DDoS", 7.5, "high"),
    ("CVE-2023-36802", "Microsoft Streaming Service Privilege Escalation", 7.8, "high"),
    ("CVE-2023-35311", "Microsoft Outlook Security Feature Bypass", 8.8, "high"),
    ("CVE-2023-32315", "OpenFire Admin Console Authentication Bypass", 7.5, "high"),
    ("CVE-2023-27350", "PaperCut NG/MF Authentication Bypass", 9.8, "critical"),
    ("CVE-2023-25717", "Ruckus Wireless Admin RCE", 9.8, "critical"),
    ("CVE-2022-27518", "Citrix ADC and Gateway RCE", 9.8, "critical"),
    ("CVE-2022-1388", "F5 BIG-IP iControl REST Authentication Bypass", 9.8, "critical"),
    # Medium severity
    ("CVE-2023-44981", "Apache ZooKeeper Authorization Bypass", 6.5, "medium"),
    ("CVE-2023-39417", "PostgreSQL SQL Injection", 6.5, "medium"),
    ("CVE-2023-38408", "OpenSSH Agent Forwarding RCE", 6.5, "medium"),
    ("CVE-2023-37920", "Certifi Removal of e-Tugra Root Certificate", 6.5, "medium"),
    ("CVE-2023-35116", "Jackson Databind Denial of Service", 5.9, "medium"),
    ("CVE-2023-33201", "Bouncy Castle Certificate Validation Bypass", 5.3, "medium"),
    ("CVE-2023-3089", "OpenShift Container Platform Incomplete Cleanup", 5.5, "medium"),
    ("CVE-2023-2976", "Google Guava Temp Directory Creation Issue", 5.5, "medium"),
    ("CVE-2023-28709", "Apache Tomcat DoS", 5.3, "medium"),
    ("CVE-2023-26049", "Eclipse Jetty Cookie Parsing Issue", 5.3, "medium"),
    ("CVE-2023-24998", "Apache Commons FileUpload DoS", 6.5, "medium"),
    ("CVE-2023-20863", "Spring Expression DoS", 5.3, "medium"),
    ("CVE-2023-1370", "Json-smart Denial of Service", 5.9, "medium"),
    ("CVE-2022-45688", "JSON-java Denial of Service", 5.9, "medium"),
    ("CVE-2022-45685", "Jettison Denial of Service", 5.9, "medium"),
    # Low severity
    ("CVE-2023-45133", "Babel Sandbox Escape", 3.7, "low"),
    ("CVE-2023-42503", "Apache Commons Compress Denial of Service", 3.3, "low"),
    ("CVE-2023-40167", "Eclipse Jetty HTTP Response Smuggling", 3.7, "low"),
    ("CVE-2023-36479", "Eclipse Jetty CGI Servlet Information Disclosure", 3.5, "low"),
    ("CVE-2023-34455", "Snappy Decompress Memory Issue", 3.3, "low"),
    ("CVE-2023-33953", "gRPC Denial of Service", 3.7, "low"),
    ("CVE-2023-2650", "OpenSSL DoS", 3.7, "low"),
    ("CVE-2023-0286", "OpenSSL X.400 Address Type Confusion", 3.7, "low"),
    ("CVE-2022-42004", "Jackson Databind Resource Exhaustion", 3.7, "low"),
    ("CVE-2022-42003", "Jackson Databind Resource Exhaustion", 3.7, "low"),
]

HOSTS = [
    ("webserver01.internal.corp", "192.168.1.10", "CentOS 7.9", "Web Servers"),
    ("webserver02.internal.corp", "192.168.1.11", "CentOS 7.9", "Web Servers"),
    ("webserver03.internal.corp", "192.168.1.12", "Ubuntu 22.04", "Web Servers"),
    ("api-prod-01.internal.corp", "192.168.1.20", "Ubuntu 20.04", "API Servers"),
    ("api-prod-02.internal.corp", "192.168.1.21", "Ubuntu 20.04", "API Servers"),
    ("api-staging-01.internal.corp", "192.168.1.25", "Ubuntu 20.04", "API Servers"),
    ("db-prod-01.internal.corp", "192.168.1.30", "RHEL 8.6", "Database Servers"),
    ("db-prod-02.internal.corp", "192.168.1.31", "RHEL 8.6", "Database Servers"),
    ("db-replica-01.internal.corp", "192.168.1.35", "RHEL 8.6", "Database Servers"),
    ("mail-server.internal.corp", "192.168.1.40", "Windows Server 2019", "Mail Servers"),
    ("exchange01.internal.corp", "192.168.1.41", "Windows Server 2019", "Mail Servers"),
    ("fileserver01.internal.corp", "192.168.1.50", "Windows Server 2016", "File Servers"),
    ("fileserver02.internal.corp", "192.168.1.51", "Windows Server 2016", "File Servers"),
    ("dc01.internal.corp", "192.168.1.60", "Windows Server 2019", "Domain Controllers"),
    ("dc02.internal.corp", "192.168.1.61", "Windows Server 2019", "Domain Controllers"),
    ("vpn-gateway.internal.corp", "192.168.1.70", "Cisco IOS 15.9", "Network Devices"),
    ("fw-external.internal.corp", "192.168.1.71", "Fortinet FortiOS 7.0", "Firewalls"),
    ("fw-internal.internal.corp", "192.168.1.72", "Palo Alto PAN-OS 10.2", "Firewalls"),
    ("jump-server.internal.corp", "192.168.1.80", "Windows Server 2022", "Jump Servers"),
    ("monitoring.internal.corp", "192.168.1.90", "Ubuntu 22.04", "Monitoring"),
    ("jenkins.internal.corp", "192.168.1.100", "Ubuntu 20.04", "CI/CD"),
    ("gitlab.internal.corp", "192.168.1.101", "Ubuntu 20.04", "CI/CD"),
    ("artifactory.internal.corp", "192.168.1.102", "RHEL 8.6", "CI/CD"),
    ("k8s-master-01.internal.corp", "192.168.1.110", "Ubuntu 22.04", "Kubernetes"),
    ("k8s-worker-01.internal.corp", "192.168.1.111", "Ubuntu 22.04", "Kubernetes"),
    ("k8s-worker-02.internal.corp", "192.168.1.112", "Ubuntu 22.04", "Kubernetes"),
    ("k8s-worker-03.internal.corp", "192.168.1.113", "Ubuntu 22.04", "Kubernetes"),
    ("splunk.internal.corp", "192.168.1.120", "RHEL 8.6", "SIEM"),
    ("backup-server.internal.corp", "192.168.1.130", "Windows Server 2019", "Backup"),
    ("dev-workstation-01", "192.168.2.10", "Windows 11", "Workstations"),
]

PORTS = [22, 80, 443, 445, 1433, 3306, 5432, 8080, 8443, 9090, 3389, 5985]
PROTOCOLS = ["tcp", "udp"]
SERVICES = ["ssh", "http", "https", "smb", "mssql", "mysql", "postgresql", "http-proxy", "https-alt", "rdp", "winrm"]


def generate_qualys_week1():
    """Generate baseline Qualys report with 100 vulnerabilities."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerability Report"

    # Qualys column headers
    headers = [
        "IP Address", "Hostname", "OS", "QID", "Title", "Severity",
        "CVE ID", "CVSS Base Score", "QDS", "Threat", "Impact",
        "Solution", "Port", "Protocol", "Service", "First Detected",
        "Last Detected", "Asset Group"
    ]
    ws.append(headers)

    vulnerabilities = []
    random.seed(42)  # For reproducibility

    # Generate 100 vulnerabilities
    for i in range(100):
        cve_data = random.choice(CVES)
        host_data = random.choice(HOSTS)
        port = random.choice(PORTS)

        # Map severity to Qualys scale (1-5)
        severity_map = {"critical": 5, "high": 4, "medium": 3, "low": 2, "info": 1}
        qualys_severity = severity_map.get(cve_data[3], 3)

        qid = 10000 + i
        qds = random.randint(60, 100) if qualys_severity >= 4 else random.randint(20, 70)

        first_detected = datetime(2025, 12, 15) + timedelta(days=random.randint(0, 20))
        last_detected = datetime(2026, 1, 3)

        row = [
            host_data[1],  # IP Address
            host_data[0],  # Hostname
            host_data[2],  # OS
            qid,  # QID
            cve_data[1],  # Title
            qualys_severity,  # Severity
            cve_data[0],  # CVE ID
            cve_data[2],  # CVSS Base Score
            qds,  # QDS
            f"This vulnerability allows attackers to {cve_data[1].lower().split()[-1]} the system.",  # Threat
            "Successful exploitation could lead to system compromise.",  # Impact
            f"Apply vendor patches. Update to the latest version.",  # Solution
            port,  # Port
            random.choice(PROTOCOLS),  # Protocol
            random.choice(SERVICES),  # Service
            first_detected.strftime("%Y-%m-%d %H:%M:%S"),  # First Detected
            last_detected.strftime("%Y-%m-%d %H:%M:%S"),  # Last Detected
            host_data[3],  # Asset Group
        ]
        ws.append(row)
        vulnerabilities.append((host_data[0], cve_data[0], qid))

    return wb, vulnerabilities


def generate_qualys_week2(week1_vulns):
    """Generate Week 2 Qualys report: 90 from week 1 + 10 new."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerability Report"

    headers = [
        "IP Address", "Hostname", "OS", "QID", "Title", "Severity",
        "CVE ID", "CVSS Base Score", "QDS", "Threat", "Impact",
        "Solution", "Port", "Protocol", "Service", "First Detected",
        "Last Detected", "Asset Group"
    ]
    ws.append(headers)

    random.seed(43)

    # Keep 90 from week 1 (remove 10 random ones)
    kept_vulns = random.sample(week1_vulns, 90)

    for hostname, cve, qid in kept_vulns:
        host_data = next((h for h in HOSTS if h[0] == hostname), HOSTS[0])
        cve_data = next((c for c in CVES if c[0] == cve), CVES[0])

        severity_map = {"critical": 5, "high": 4, "medium": 3, "low": 2, "info": 1}
        qualys_severity = severity_map.get(cve_data[3], 3)
        qds = random.randint(60, 100) if qualys_severity >= 4 else random.randint(20, 70)

        first_detected = datetime(2025, 12, 15) + timedelta(days=random.randint(0, 20))
        last_detected = datetime(2026, 1, 10)

        row = [
            host_data[1], host_data[0], host_data[2], qid, cve_data[1],
            qualys_severity, cve_data[0], cve_data[2], qds,
            f"This vulnerability allows attackers to compromise the system.",
            "Successful exploitation could lead to system compromise.",
            "Apply vendor patches.", random.choice(PORTS),
            random.choice(PROTOCOLS), random.choice(SERVICES),
            first_detected.strftime("%Y-%m-%d %H:%M:%S"),
            last_detected.strftime("%Y-%m-%d %H:%M:%S"),
            host_data[3],
        ]
        ws.append(row)

    # Add 10 new vulnerabilities
    new_vulns = []
    for i in range(10):
        cve_data = random.choice(CVES)
        host_data = random.choice(HOSTS)
        qid = 20000 + i

        severity_map = {"critical": 5, "high": 4, "medium": 3, "low": 2, "info": 1}
        qualys_severity = severity_map.get(cve_data[3], 3)
        qds = random.randint(60, 100) if qualys_severity >= 4 else random.randint(20, 70)

        first_detected = datetime(2026, 1, 8)
        last_detected = datetime(2026, 1, 10)

        row = [
            host_data[1], host_data[0], host_data[2], qid, cve_data[1],
            qualys_severity, cve_data[0], cve_data[2], qds,
            f"This vulnerability allows attackers to compromise the system.",
            "Successful exploitation could lead to system compromise.",
            "Apply vendor patches.", random.choice(PORTS),
            random.choice(PROTOCOLS), random.choice(SERVICES),
            first_detected.strftime("%Y-%m-%d %H:%M:%S"),
            last_detected.strftime("%Y-%m-%d %H:%M:%S"),
            host_data[3],
        ]
        ws.append(row)
        new_vulns.append((host_data[0], cve_data[0], qid))

    return wb, kept_vulns, new_vulns


def generate_qualys_week3(week1_kept, week2_new):
    """Generate Week 3 Qualys report: 72 from week 1 + 10 from week 2 + 5 new."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerability Report"

    headers = [
        "IP Address", "Hostname", "OS", "QID", "Title", "Severity",
        "CVE ID", "CVSS Base Score", "QDS", "Threat", "Impact",
        "Solution", "Port", "Protocol", "Service", "First Detected",
        "Last Detected", "Asset Group"
    ]
    ws.append(headers)

    random.seed(44)

    # Keep 72 from week 1's remaining 90 (close 20% = 18)
    week1_remaining = random.sample(week1_kept, 72)

    for hostname, cve, qid in week1_remaining + week2_new:
        host_data = next((h for h in HOSTS if h[0] == hostname), HOSTS[0])
        cve_data = next((c for c in CVES if c[0] == cve), CVES[0])

        severity_map = {"critical": 5, "high": 4, "medium": 3, "low": 2, "info": 1}
        qualys_severity = severity_map.get(cve_data[3], 3)
        qds = random.randint(60, 100) if qualys_severity >= 4 else random.randint(20, 70)

        first_detected = datetime(2025, 12, 15) + timedelta(days=random.randint(0, 20))
        last_detected = datetime(2026, 1, 17)

        row = [
            host_data[1], host_data[0], host_data[2], qid, cve_data[1],
            qualys_severity, cve_data[0], cve_data[2], qds,
            f"This vulnerability allows attackers to compromise the system.",
            "Successful exploitation could lead to system compromise.",
            "Apply vendor patches.", random.choice(PORTS),
            random.choice(PROTOCOLS), random.choice(SERVICES),
            first_detected.strftime("%Y-%m-%d %H:%M:%S"),
            last_detected.strftime("%Y-%m-%d %H:%M:%S"),
            host_data[3],
        ]
        ws.append(row)

    # Add 5 new vulnerabilities
    for i in range(5):
        cve_data = random.choice(CVES)
        host_data = random.choice(HOSTS)
        qid = 30000 + i

        severity_map = {"critical": 5, "high": 4, "medium": 3, "low": 2, "info": 1}
        qualys_severity = severity_map.get(cve_data[3], 3)
        qds = random.randint(60, 100) if qualys_severity >= 4 else random.randint(20, 70)

        first_detected = datetime(2026, 1, 15)
        last_detected = datetime(2026, 1, 17)

        row = [
            host_data[1], host_data[0], host_data[2], qid, cve_data[1],
            qualys_severity, cve_data[0], cve_data[2], qds,
            f"This vulnerability allows attackers to compromise the system.",
            "Successful exploitation could lead to system compromise.",
            "Apply vendor patches.", random.choice(PORTS),
            random.choice(PROTOCOLS), random.choice(SERVICES),
            first_detected.strftime("%Y-%m-%d %H:%M:%S"),
            last_detected.strftime("%Y-%m-%d %H:%M:%S"),
            host_data[3],
        ]
        ws.append(row)

    return wb


def generate_tenable_week1():
    """Generate baseline Tenable report with 100 vulnerabilities."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    headers = [
        "Plugin ID", "CVE", "CVSS v3.0 Base Score", "Risk", "Host",
        "Protocol", "Port", "Name", "Synopsis", "Description",
        "Solution", "VPR", "Plugin Family", "First Discovered", "Last Seen"
    ]
    ws.append(headers)

    vulnerabilities = []
    random.seed(52)

    for i in range(100):
        cve_data = random.choice(CVES)
        host_data = random.choice(HOSTS)
        port = random.choice(PORTS)

        # Map severity to Tenable risk levels
        risk_map = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low", "info": "Info"}
        risk = risk_map.get(cve_data[3], "Medium")

        plugin_id = 100000 + i
        vpr = round(random.uniform(7.0, 10.0), 1) if cve_data[3] in ["critical", "high"] else round(random.uniform(3.0, 7.0), 1)

        first_discovered = datetime(2025, 12, 15) + timedelta(days=random.randint(0, 20))
        last_seen = datetime(2026, 1, 3)

        plugin_families = ["Web Servers", "Databases", "Windows", "Ubuntu Local Security Checks",
                          "Red Hat Local Security Checks", "Firewalls", "General"]

        row = [
            plugin_id,  # Plugin ID
            cve_data[0],  # CVE
            cve_data[2],  # CVSS v3.0 Base Score
            risk,  # Risk
            host_data[0],  # Host
            random.choice(PROTOCOLS),  # Protocol
            port,  # Port
            cve_data[1],  # Name
            f"The remote host is affected by {cve_data[1].lower()}.",  # Synopsis
            f"The remote host is running software affected by {cve_data[0]}. {cve_data[1]}.",  # Description
            "Apply the latest vendor patches.",  # Solution
            vpr,  # VPR
            random.choice(plugin_families),  # Plugin Family
            first_discovered.strftime("%Y-%m-%d %H:%M:%S"),  # First Discovered
            last_seen.strftime("%Y-%m-%d %H:%M:%S"),  # Last Seen
        ]
        ws.append(row)
        vulnerabilities.append((host_data[0], cve_data[0], plugin_id))

    return wb, vulnerabilities


def generate_tenable_week2(week1_vulns):
    """Generate Week 2 Tenable report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    headers = [
        "Plugin ID", "CVE", "CVSS v3.0 Base Score", "Risk", "Host",
        "Protocol", "Port", "Name", "Synopsis", "Description",
        "Solution", "VPR", "Plugin Family", "First Discovered", "Last Seen"
    ]
    ws.append(headers)

    random.seed(53)
    kept_vulns = random.sample(week1_vulns, 90)

    plugin_families = ["Web Servers", "Databases", "Windows", "Ubuntu Local Security Checks",
                      "Red Hat Local Security Checks", "Firewalls", "General"]

    for hostname, cve, plugin_id in kept_vulns:
        host_data = next((h for h in HOSTS if h[0] == hostname), HOSTS[0])
        cve_data = next((c for c in CVES if c[0] == cve), CVES[0])

        risk_map = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low", "info": "Info"}
        risk = risk_map.get(cve_data[3], "Medium")
        vpr = round(random.uniform(7.0, 10.0), 1) if cve_data[3] in ["critical", "high"] else round(random.uniform(3.0, 7.0), 1)

        first_discovered = datetime(2025, 12, 15) + timedelta(days=random.randint(0, 20))
        last_seen = datetime(2026, 1, 10)

        row = [
            plugin_id, cve_data[0], cve_data[2], risk, host_data[0],
            random.choice(PROTOCOLS), random.choice(PORTS), cve_data[1],
            f"The remote host is affected by {cve_data[1].lower()}.",
            f"The remote host is running software affected by {cve_data[0]}.",
            "Apply the latest vendor patches.", vpr,
            random.choice(plugin_families),
            first_discovered.strftime("%Y-%m-%d %H:%M:%S"),
            last_seen.strftime("%Y-%m-%d %H:%M:%S"),
        ]
        ws.append(row)

    # Add 10 new
    new_vulns = []
    for i in range(10):
        cve_data = random.choice(CVES)
        host_data = random.choice(HOSTS)
        plugin_id = 200000 + i

        risk_map = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low", "info": "Info"}
        risk = risk_map.get(cve_data[3], "Medium")
        vpr = round(random.uniform(7.0, 10.0), 1) if cve_data[3] in ["critical", "high"] else round(random.uniform(3.0, 7.0), 1)

        row = [
            plugin_id, cve_data[0], cve_data[2], risk, host_data[0],
            random.choice(PROTOCOLS), random.choice(PORTS), cve_data[1],
            f"The remote host is affected by {cve_data[1].lower()}.",
            f"The remote host is running software affected by {cve_data[0]}.",
            "Apply the latest vendor patches.", vpr,
            random.choice(plugin_families),
            datetime(2026, 1, 8).strftime("%Y-%m-%d %H:%M:%S"),
            datetime(2026, 1, 10).strftime("%Y-%m-%d %H:%M:%S"),
        ]
        ws.append(row)
        new_vulns.append((host_data[0], cve_data[0], plugin_id))

    return wb, kept_vulns, new_vulns


def generate_tenable_week3(week1_kept, week2_new):
    """Generate Week 3 Tenable report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    headers = [
        "Plugin ID", "CVE", "CVSS v3.0 Base Score", "Risk", "Host",
        "Protocol", "Port", "Name", "Synopsis", "Description",
        "Solution", "VPR", "Plugin Family", "First Discovered", "Last Seen"
    ]
    ws.append(headers)

    random.seed(54)
    week1_remaining = random.sample(week1_kept, 72)

    plugin_families = ["Web Servers", "Databases", "Windows", "Ubuntu Local Security Checks",
                      "Red Hat Local Security Checks", "Firewalls", "General"]

    for hostname, cve, plugin_id in week1_remaining + week2_new:
        host_data = next((h for h in HOSTS if h[0] == hostname), HOSTS[0])
        cve_data = next((c for c in CVES if c[0] == cve), CVES[0])

        risk_map = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low", "info": "Info"}
        risk = risk_map.get(cve_data[3], "Medium")
        vpr = round(random.uniform(7.0, 10.0), 1) if cve_data[3] in ["critical", "high"] else round(random.uniform(3.0, 7.0), 1)

        first_discovered = datetime(2025, 12, 15) + timedelta(days=random.randint(0, 20))
        last_seen = datetime(2026, 1, 17)

        row = [
            plugin_id, cve_data[0], cve_data[2], risk, host_data[0],
            random.choice(PROTOCOLS), random.choice(PORTS), cve_data[1],
            f"The remote host is affected by {cve_data[1].lower()}.",
            f"The remote host is running software affected by {cve_data[0]}.",
            "Apply the latest vendor patches.", vpr,
            random.choice(plugin_families),
            first_discovered.strftime("%Y-%m-%d %H:%M:%S"),
            last_seen.strftime("%Y-%m-%d %H:%M:%S"),
        ]
        ws.append(row)

    # Add 5 new
    for i in range(5):
        cve_data = random.choice(CVES)
        host_data = random.choice(HOSTS)
        plugin_id = 300000 + i

        risk_map = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low", "info": "Info"}
        risk = risk_map.get(cve_data[3], "Medium")
        vpr = round(random.uniform(7.0, 10.0), 1) if cve_data[3] in ["critical", "high"] else round(random.uniform(3.0, 7.0), 1)

        row = [
            plugin_id, cve_data[0], cve_data[2], risk, host_data[0],
            random.choice(PROTOCOLS), random.choice(PORTS), cve_data[1],
            f"The remote host is affected by {cve_data[1].lower()}.",
            f"The remote host is running software affected by {cve_data[0]}.",
            "Apply the latest vendor patches.", vpr,
            random.choice(plugin_families),
            datetime(2026, 1, 15).strftime("%Y-%m-%d %H:%M:%S"),
            datetime(2026, 1, 17).strftime("%Y-%m-%d %H:%M:%S"),
        ]
        ws.append(row)

    return wb


if __name__ == "__main__":
    import os

    script_dir = os.path.dirname(os.path.abspath(__file__))

    print("Generating Qualys reports...")
    # Qualys Week 1
    wb1, vulns1 = generate_qualys_week1()
    wb1.save(os.path.join(script_dir, "qualys_week1.xlsx"))
    print(f"  Week 1: 100 vulnerabilities")

    # Qualys Week 2
    wb2, kept2, new2 = generate_qualys_week2(vulns1)
    wb2.save(os.path.join(script_dir, "qualys_week2.xlsx"))
    print(f"  Week 2: 90 existing + 10 new = 100 vulnerabilities")

    # Qualys Week 3
    wb3 = generate_qualys_week3(kept2, new2)
    wb3.save(os.path.join(script_dir, "qualys_week3.xlsx"))
    print(f"  Week 3: 72 from week1 + 10 from week2 + 5 new = 87 vulnerabilities")

    print("\nGenerating Tenable reports...")
    # Tenable Week 1
    wb1t, vulns1t = generate_tenable_week1()
    wb1t.save(os.path.join(script_dir, "tenable_week1.xlsx"))
    print(f"  Week 1: 100 vulnerabilities")

    # Tenable Week 2
    wb2t, kept2t, new2t = generate_tenable_week2(vulns1t)
    wb2t.save(os.path.join(script_dir, "tenable_week2.xlsx"))
    print(f"  Week 2: 90 existing + 10 new = 100 vulnerabilities")

    # Tenable Week 3
    wb3t = generate_tenable_week3(kept2t, new2t)
    wb3t.save(os.path.join(script_dir, "tenable_week3.xlsx"))
    print(f"  Week 3: 72 from week1 + 10 from week2 + 5 new = 87 vulnerabilities")

    print("\nDone! Generated 6 test files in", script_dir)
